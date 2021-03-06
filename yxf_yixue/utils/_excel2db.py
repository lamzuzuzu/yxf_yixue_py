#!/usr/bin/python3
# -*- coding: utf-8 -*-
import os
import datetime
import openpyxl
import openpyxl.utils
from ._db import Db

"""
openpyxl说明：

    表格操作：
    wb = Workbook()
        ——在内存中建立表格对象
    ws = wb.active
        ——获取当前活跃的sheet，默认为第一个
    sheets = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheets[0])
        ——获取所有sheet，然后选择特定的sheet
    wb.save("path")
        ——保存文件
    wb = load_workbook("path")
        ——读取已存在表格文件

    多条数据操作：
    rows = ws.rows
    columns = ws.columns
        ——获取表格所有行、列，均可迭代
    cell_range = ws['A1':'C2']
    colC = ws['C']
    col_range = ws['C:D']
    row10 = ws[10]
    row_range = ws[5:10]
        ——获取一部分内容

    单条数据操作：
    ws.cell(row=1, column=1).value = 6
        ——单元格数字寻址，赋值，起始于1
    ws.cell("B1").value = 7
        ——单元格字母寻址，赋值，起始于1
    get_column_letter(col)
        ——获取单元格字母地址
    ws.append(["我","你","她"])
        ——在末尾插入行
    ws['A1'].value
    ws['A1] = 4
        ——单元格寻址，赋值
"""


class Excel2Db:
    def __init__(self, dbname):
        self.db = Db(dbname)
        self.dbname = dbname

    # 读取Excel
    def read_excel(self, file_path, file_name):
        realpath = os.path.join(file_path, file_name)
        wb_origin = openpyxl.load_workbook(realpath)
        ws_origin = wb_origin.active  # excel对象
        table_colname = []
        table = []
        for i, row in enumerate(ws_origin):  # 逐行解析
            if i == 0:  # 标题行
                for col in ws_origin[i + 1]:
                    table_colname.append(col.value)
            else:
                table_col = []
                for col in ws_origin[i + 1]:  # 循环加载每一数据行的列值，第一个空行也会被加载（随后会被for循环外部的判断退出）
                    if col.value:  # 正常值
                        if isinstance(col.value, datetime.datetime):  # 如果是日期时间格式则转化为字符串，独立逻辑
                            table_col.append(col.value.strftime('%Y-%m-%d'))
                        else:
                            table_col.append(col.value)
                    else:
                        if self.dbname == 'utils.db':
                            table_col.append(col.value)
                        else:
                            # 曾经出现过奇怪的BUG，数字0被认为是空值？后来BUG又自己好了？
                            # 为以防万一，真正的空值在原表中用字符串'None'填充，为了区别，原表不能有空值
                            # table_col.append(0)
                            table_col.append(col.value)
                if table_col[0] is None or table_col[0] is 'None':  # 遇到空行退出（因为表格后面有自己加的注释）
                    break
                table.append(table_col)
        return file_name.split('.')[0], table_colname, table

    # 搜索所有的excel，根据首行自动添加数据列，原封不动地存入数据库
    def transform2db(self, excelpath):
        file_list = os.listdir(excelpath)
        for file in file_list:
            tablename, table_colname, table = self.read_excel(excelpath, file)
            sqlstr_create = "create table [{0}] (".format(tablename)
            for col in table_colname:
                if col == "序号":
                    sqlstr_create += "[{0}] text primary key not null,".format(col)
                else:
                    sqlstr_create += "[{0}] text not null,".format(col)
            # 直接笨办法：删除旧表，创建新表
            self.db.execute("drop table if exists [{0}];".format(tablename))
            self.db.commit()
            print("drop table if exists [{0}];".format(tablename))
            self.db.execute(sqlstr_create[:-1]+");")
            self.db.commit()
            print(sqlstr_create[:-1] + ");")
            # 存入数据
            table_colnamestr = ""
            for i in table_colname:
                table_colnamestr += "["+i+"],"
            for row in table:
                sqlstr_insert = "insert into [{0}]({1})values({2});".format(tablename,table_colnamestr[:-1],str(row).lstrip('[').rstrip(']'))
                print(sqlstr_insert)
                self.db.execute(sqlstr_insert)
            self.db.commit()
